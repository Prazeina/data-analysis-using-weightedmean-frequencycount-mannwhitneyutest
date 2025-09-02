import pandas as pd
from scipy.stats import mannwhitneyu,mode
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import dataframe_image as dfi

# Load the CSV files
graduates_data_new = pd.read_csv('/Users/prazeina/Documents/repos/RS/2024/mann-whitney/testFiles/female_confidence_frequency.csv', index_col=0)
undergraduates_data_new = pd.read_csv('/Users/prazeina/Documents/repos/RS/2024/mann-whitney/testFiles/male_confidence_frequency.csv', index_col=0)

# Define the ranks as Excellent, Very Good, Good, Fair, Poor
rank_mapping = {
    'Excellent': 1,
    'Very Good': 2,
    'Good': 3,
    'Fair': 4,
    'Poor': 5
}

# Function to expand frequency counts into individual ranks using the new ranking system
def expand_ranks_new(freq_series):
    expanded_ranks = []
    for label, count in freq_series.items():
        expanded_ranks.extend([rank_mapping[label]] * int(count))
    return expanded_ranks

# Initialize lists to store results
skills = graduates_data_new.columns
results = {
    'Skill': [],
    'Median_F': [],
    'Median_M': [],
    'Mode_Female': [],
    'Mode_Male': [],
    'Mean_F': [],
    'Mean_M': [],
    'Sample_Size_F': [],
    'Sample_Size_M': [],
    'U_Statistic': [],
    'P_Value': [],
    'Preferred_Group': []
}

# Process each skill for the new data
for skill in skills:
    # Get frequency counts for this skill
    grad_freq = graduates_data_new[skill]
    undergrad_freq = undergraduates_data_new[skill]
    
    # Expand frequencies into individual ranks
    grad_ranks = expand_ranks_new(grad_freq)
    undergrad_ranks = expand_ranks_new(undergrad_freq)
    
    # Calculate medians
    median_grad = np.median(grad_ranks) if grad_ranks else np.nan
    median_undergrad = np.median(undergrad_ranks) if undergrad_ranks else np.nan
    mean_grad = np.mean(grad_ranks) if grad_ranks else np.nan
    mean_undergrad = np.mean(undergrad_ranks) if undergrad_ranks else np.nan
    # Assuming grad_ranks and undergrad_ranks are defined earlier
    mode_grad = mode(grad_ranks).mode if grad_ranks else np.nan
    mode_undergrad = mode(undergrad_ranks).mode if undergrad_ranks else np.nan

    # Calculate sample sizes
    n_grad = len(grad_ranks)
    n_undergrad = len(undergrad_ranks)
    
    # Perform Mann-Whitney U test
    if n_grad > 0 and n_undergrad > 0:  # Ensure both groups have data
        u_stat, p_val = mannwhitneyu(grad_ranks, undergrad_ranks, alternative='two-sided')
    else:
        u_stat, p_val = np.nan, np.nan

    # Determine preferred group
    if mean_grad < mean_undergrad:
        preferred = 'F'
    elif mean_grad > mean_undergrad:
        preferred = 'M'
    else:
        preferred = 'Equal'
    
    # Store results
    results['Skill'].append(skill)
    results['Median_F'].append(median_grad)
    results['Median_M'].append(median_undergrad)
    results['Mean_F'].append(mean_grad)
    results['Mean_M'].append(mean_undergrad)
    results['Mode_Female'].append(mode_grad)
    results['Mode_Male'].append(mode_undergrad)
    results['Sample_Size_F'].append(n_grad)
    results['Sample_Size_M'].append(n_undergrad)
    results['U_Statistic'].append(u_stat)
    results['P_Value'].append(p_val)
    results['Preferred_Group'].append(preferred)

# Create a DataFrame from results
results_df = pd.DataFrame(results)
# results['Preferred_Group'].append(preferred)
results_df = results_df.sort_values(by='U_Statistic', ascending=False)
# Save to CSV
results_df.to_csv('/Users/prazeina/Documents/repos/RS/2024/mann-whitney/result/gender_confidence_mannWhitneyU_results_all.csv', index=False)

# ------------------------------------------------------------------------------------------------------------------------------
import pandas as pd
import dataframe_image as dfi

# Define header rows for Excel
header_row1 = ['', 'Female', '', 'Male', '', '']
header_row2 = ['Skill', 'Median', 'N', 'Median', 'N', 'P-Value']

# Create the new DataFrame without the U-Statistic column
output_df = pd.DataFrame({
    'Skill': results_df['Skill'],
    'Female_Median': results_df['Median_F'],
    'Female_N': results_df['Sample_Size_F'],
    'Male_Median': results_df['Median_M'],
    'Male_N': results_df['Sample_Size_M'],
    'P-Value': results_df['P_Value']
})

# Sort by P-Value (instead of U_Statistic) for significance
output_df = output_df.sort_values(by='P-Value', ascending=True)

# Create Excel file
wb = Workbook()
ws = wb.active
ws.title = "Results"

# Write the header rows
ws.append(header_row1)
ws.append(header_row2)

# Write the data rows
for row in dataframe_to_rows(output_df, index=False, header=False):
    ws.append(row)

# Merging cells under "Female" and "Male" headers
ws.merge_cells('B1:C1')  # Merge Female columns (Median and N)
ws.merge_cells('D1:E1')  # Merge Male columns (Median and N)

# Apply center alignment for merged cells
ws['B1'].alignment = Alignment(horizontal='center')
ws['D1'].alignment = Alignment(horizontal='center')

# Save the Excel file
excel_output_file_path = '/Users/prazeina/Documents/repos/RS/2024/mann-whitney/result/gender_confidence_mannWhitneyU_results.xlsx'
wb.save(excel_output_file_path)

# --- Create styled DataFrame for image export ---
# Read the Excel file
df = pd.read_excel(
    "/Users/prazeina/Documents/repos/RS/2024/mann-whitney/result/gender_confidence_mannWhitneyU_results.xlsx",
    sheet_name="Results",
    header=None
)

# Set the first two rows as MultiIndex header
df.columns = pd.MultiIndex.from_arrays([df.iloc[0], df.iloc[1]])
df = df.drop([0, 1]).reset_index(drop=True)

# Assign clean column names without U-Statistic
new_columns = pd.MultiIndex.from_tuples([
    ('', 'Skill'),
    ('Female', 'Median'),
    ('Female', 'N'),
    ('Male', 'Median'),
    ('Male', 'N'),
    ('', 'p-value')
])
df.columns = new_columns

# Function to apply border styling based on p-value
def highlight_pval(val):
    try:
        val = float(val)
        if val < 0.05:
            return 'border: 2px solid red'
        elif val < 0.1:
            return 'border: 2px solid green'
        else:
            return ''
    except:
        return ''

# Style the DataFrame
df_styled = df.style \
    .hide(axis='index') \
    .format({
        ('', 'p-value'): lambda x: f"{x:.3f}".lstrip("0"),
        ('Female', 'Median'): lambda x: f"{float(x):.1f}".rstrip("0").rstrip(".").lstrip("0") if "." in f"{float(x):.1f}" else f"{int(x)}",
        ('Female', 'N'): "{:.0f}",
        ('Male', 'Median'): lambda x: f"{float(x):.1f}".rstrip("0").rstrip(".").lstrip("0") if "." in f"{float(x):.1f}" else f"{int(x)}",
        ('Male', 'N'): "{:.0f}"
    }) \
    .applymap(highlight_pval, subset=[('', 'p-value')]) \
    .set_table_styles([
        {'selector': 'tbody tr:nth-child(even)', 'props': [('background-color', '#f9f9f9')]}
    ])

# Export as image
dfi.export(df_styled, "/Users/prazeina/Documents/repos/RS/2024/mann-whitney/result/2024 Confidence Gender.png", dpi=300)