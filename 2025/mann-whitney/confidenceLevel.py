import pandas as pd
from scipy.stats import mannwhitneyu,mode
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
import pandas as pd
import dataframe_image as dfi

# Load the CSV files
graduates_data_new = pd.read_csv('/Users/prazeina/Documents/repos/RS/2025/mann-whitney/testFiles/grad_processed_mean_values.csv', index_col=0)
undergraduates_data_new = pd.read_csv('/Users/prazeina/Documents/repos/RS/2025/mann-whitney/testFiles/undergrad_processed_mean_values.csv', index_col=0)

# Define the ranks as Excellent, Very Good, Good, Fair, Poor
rank_mapping = {
    'Excellent': 1,
    'Very Good': 2,
    'Good': 3,
    'Fair': 4,
    'Poor': 5
}

# Function to expand frequency counts into individual ranks using the new ranking system
def expand_ranks_new(freq_series):
    expanded_ranks = []
    for label, count in freq_series.items():
        expanded_ranks.extend([rank_mapping[label]] * int(count))
    return expanded_ranks

# Initialize lists to store results
skills = graduates_data_new.columns
results = {
    'Skill': [],
    'Median_Graduates': [],
    'Median_Undergraduates': [],
    'Mode_Graduates': [],
    'Mode_Undergraduates': [],
    'Mean_Graduates': [],
    'Mean_Undergraduates': [],
    'Sample_Size_Graduates': [],
    'Sample_Size_Undergraduates': [],
    'U_Statistic': [],
    'P_Value': [],
    'Preferred_Group': []
}

# Process each skill for the new data
for skill in skills:
    # Get frequency counts for this skill
    grad_freq = graduates_data_new[skill]
    undergrad_freq = undergraduates_data_new[skill]
    
    # Expand frequencies into individual ranks
    grad_ranks = expand_ranks_new(grad_freq)
    undergrad_ranks = expand_ranks_new(undergrad_freq)
    
    # Calculate medians
    median_grad = np.median(grad_ranks) if grad_ranks else np.nan
    median_undergrad = np.median(undergrad_ranks) if undergrad_ranks else np.nan
    mean_grad = np.mean(grad_ranks) if grad_ranks else np.nan
    mean_undergrad = np.mean(undergrad_ranks) if undergrad_ranks else np.nan
     # Assuming grad_ranks and undergrad_ranks are defined earlier
    mode_grad = mode(grad_ranks).mode if grad_ranks else np.nan
    mode_undergrad = mode(undergrad_ranks).mode if undergrad_ranks else np.nan

    # Calculate sample sizes
    n_grad = len(grad_ranks)
    n_undergrad = len(undergrad_ranks)
    
    # Perform Mann-Whitney U test
    if n_grad > 0 and n_undergrad > 0:  # Ensure both groups have data
        u_stat, p_val = mannwhitneyu(grad_ranks, undergrad_ranks, alternative='two-sided')
    else:
        u_stat, p_val = np.nan, np.nan

    # Determine preferred group
    if mean_grad < mean_undergrad:
        preferred = 'Graduates'
    elif mean_grad > mean_undergrad:
        preferred = 'Undergraduates'
    else:
        preferred = 'Equal'
    
    results['Skill'].append(skill)
    results['Median_Graduates'].append(median_grad)
    results['Median_Undergraduates'].append(median_undergrad)
    results['Mode_Graduates'].append(mode_grad)
    results['Mode_Undergraduates'].append(mode_undergrad)
    results['Mean_Graduates'].append(mean_grad)
    results['Mean_Undergraduates'].append(mean_undergrad)
    results['Sample_Size_Graduates'].append(n_grad)
    results['Sample_Size_Undergraduates'].append(n_undergrad)
    results['U_Statistic'].append(u_stat)
    results['P_Value'].append(p_val)
    results['Preferred_Group'].append(preferred)

# Create a DataFrame from results
results_df = pd.DataFrame(results)
# results['Preferred_Group'].append(preferred)
results_df = results_df.sort_values(by='U_Statistic', ascending=False)
# Save to CSV
results_df.to_csv('/Users/prazeina/Documents/repos/RS/2025/mann-whitney/result/level_confidence_mannWhitneyU_results.csv', index=False)

# ----------------------------------------------------------------------------------------------------------------------------------------

header_row1 = ['', 'Graduate', '', 'Undergraduate', '', '']
header_row2 = ['Skill', 'Median', 'N', 'Median', 'N',  'P-Value']

# Create the new DataFrame with the necessary data columns
output_df = pd.DataFrame({
    'Skill': results_df['Skill'],
    'Grad_Median': results_df['Median_Graduates'],
    'Grad_N': results_df['Sample_Size_Graduates'],
    'Undergrad_Median': results_df['Median_Undergraduates'],
    'Undergrad_N': results_df['Sample_Size_Undergraduates'],
    # 'U_Statistic': results_df['U_Statistic'],
    'P-Value': results_df['P_Value']
})

# --- Save to Excel ---
# --- Sort by U_Statistic in descending order ---
# output_df = output_df.sort_values(by='U_Statistic', ascending=False)

wb = Workbook()
ws = wb.active
ws.title = "Results"

# Write the header rows
ws.append(header_row1)
ws.append(header_row2)

# Write the data rows
for row in dataframe_to_rows(output_df, index=False, header=False):
    ws.append(row)

# Merging cells under "Graduate" and "Undergraduate" headers
ws.merge_cells('B1:C1')  # Merge Graduate columns (Median and N)
ws.merge_cells('D1:E1')  # Merge Undergraduate columns (Median and N)

# Apply center alignment for merged cells
ws['B1'].alignment = Alignment(horizontal='center')
ws['D1'].alignment = Alignment(horizontal='center')

# Save the Excel file
excel_output_file_path = '/Users/prazeina/Documents/repos/RS/2025/mann-whitney/result/level_confidence_mannWhitneyU_results.xlsx'
wb.save(excel_output_file_path)
# --------------------------------------------------------------------------------

# Read the Excel file
df = pd.read_excel(
    "/Users/prazeina/Documents/repos/RS/2025/mann-whitney/result/level_confidence_mannWhitneyU_results.xlsx",
    sheet_name="Results",
    header=None
)

# Set the first two rows as MultiIndex header
df.columns = pd.MultiIndex.from_arrays([df.iloc[0], df.iloc[1]])
df = df.drop([0, 1]).reset_index(drop=True)

# Assign clean column names
new_columns = pd.MultiIndex.from_tuples([
    ('', 'Skill'),
    ('Graduate', 'Median'),
    ('Graduate', 'N'),
    ('Undergraduate', 'Median'),
    ('Undergraduate', 'N'),
    # ('', 'U-Statistic'),
    ('', 'p-value')
])
df.columns = new_columns

# Function to apply border styling based on p-value
def highlight_pval(val):
    try:
        val = float(val)
        if val < 0.05:
            return 'border: 2px solid red'
        elif val < 0.1:
            return 'border: 2px solid green'
        else:
            return ''
    except:
        return ''

df_styled = df.style \
    .hide(axis='index') \
    .format({
        ('', 'p-value'): lambda x: f"{x:.3f}".lstrip("0"),
        # ('', 'U-Statistic'): "{:.0f}",
        ('Graduate', 'Median'): lambda x: f"{float(x):.1f}".rstrip("0").rstrip(".").lstrip("0") if "." in f"{float(x):.1f}" else f"{int(x)}",
        ('Graduate', 'N'): "{:.0f}",
        ('Undergraduate', 'Median'): lambda x: f"{float(x):.1f}".rstrip("0").rstrip(".").lstrip("0") if "." in f"{float(x):.1f}" else f"{int(x)}",
        ('Undergraduate', 'N'): "{:.0f}"
    }) \
    .applymap(highlight_pval, subset=[('', 'p-value')]) \
    .set_table_styles([
        {'selector': 'tbody tr:nth-child(even)', 'props': [('background-color', '#f9f9f9')]}
    ])

# Export as image
dfi.export(df_styled, "/Users/prazeina/Documents/repos/RS/2025/mann-whitney/result/2025 Confidence Level.png", dpi=300)


# import matplotlib.pyplot as plt

# # Load the dataset
# df = pd.read_csv("/Users/prazeina/Documents/repos/RS/2025/mann-whitney/result/level_confidence_mannWhitneyU_results.csv")

# # Extract relevant columns for plotting
# skills = df['Skill']
# mean_graduates = df['Mean_Graduates']
# mean_undergraduates = df['Mean_Undergraduates']

# # Plotting
# plt.figure(figsize=(10, 6))

# # Create two bars for each skill: one for Graduates, one for Undergraduates
# bar_width = 0.35  # Width of the bars
# index = range(len(skills))  # Position of the bars

# plt.barh([i - bar_width / 2 for i in index], mean_graduates, bar_width, label='Grad', color='#00386C')
# plt.barh([i + bar_width / 2 for i in index], mean_undergraduates, bar_width, label='UGrad', color='#FFC333')

# # Add labels and title
# plt.xlabel('Weighted Average Self-Assessment')
# plt.ylabel('Skill')
# plt.title('2025 - Weighted Mean self-assessment of skills by Level\n(Lower = More Important)')
# plt.yticks(index, skills)
# plt.legend(loc='lower right')

# # Show the plot
# plt.tight_layout()
# plt.show()
