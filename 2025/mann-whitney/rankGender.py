import pandas as pd
from scipy.stats import mannwhitneyu,mode
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

# Load the CSV files
graduates_data = pd.read_csv('/Users/prazeina/Documents/repos/RS/2025/mann-whitney/testFiles/female_ranking_frequency.csv', index_col=0)
undergraduates_data = pd.read_csv('/Users/prazeina/Documents/repos/RS/2025/mann-whitney/testFiles/male_ranking_frequency.csv', index_col=0)

# Define the ranks (1 to 5)
ranks = [1, 2, 3, 4, 5]

# Function to expand frequency counts into individual ranks
def expand_ranks(freq_series):
    expanded_ranks = []
    for rank, count in zip(ranks, freq_series):
        expanded_ranks.extend([rank] * int(count))
    return expanded_ranks

# Initialize result storage
skills = graduates_data.columns
results = {
    'Skill': [],
    'Median_F': [],
    'Median_M': [],
    'Mode_Female': [],
    'Mode_Male': [],
    # 'Mean_F': [],
    # 'Mean_M': [],
    'Sample_Size_F': [],
    'Sample_Size_M': [],
    'U_Statistic': [],
    'P_Value': [],
    # 'Preferred_Group': []
}

# Process each skill
for skill in skills:
    grad_freq = graduates_data[skill]
    undergrad_freq = undergraduates_data[skill]

    grad_ranks = expand_ranks(grad_freq)
    undergrad_ranks = expand_ranks(undergrad_freq)

    # Central tendencies
    median_grad = np.median(grad_ranks) if grad_ranks else np.nan
    median_undergrad = np.median(undergrad_ranks) if undergrad_ranks else np.nan
    mean_grad = np.mean(grad_ranks) if grad_ranks else np.nan
    mean_undergrad = np.mean(undergrad_ranks) if undergrad_ranks else np.nan
 # Assuming grad_ranks and undergrad_ranks are defined earlier
    mode_grad = mode(grad_ranks).mode if grad_ranks else np.nan
    mode_undergrad = mode(undergrad_ranks).mode if undergrad_ranks else np.nan
    n_grad = len(grad_ranks)
    n_undergrad = len(undergrad_ranks)

    # Mann-Whitney U test
    if n_grad > 0 and n_undergrad > 0:
        u_stat, p_val = mannwhitneyu(grad_ranks, undergrad_ranks, alternative='two-sided')
    else:
        u_stat, p_val = np.nan, np.nan

    # Determine preferred group
    if mean_grad < mean_undergrad:
        preferred = 'F'
    elif mean_grad > mean_undergrad:
        preferred = 'M'
    else:
        preferred = 'Equal'

    # Store results
    results['Skill'].append(skill)
    results['Median_F'].append(median_grad)
    results['Median_M'].append(median_undergrad)
    results['Mode_Female'].append(mode_grad)
    results['Mode_Male'].append(mode_undergrad)
    # results['Mean_F'].append(mean_grad)
    # results['Mean_M'].append(mean_undergrad)
    results['Sample_Size_F'].append(n_grad)
    results['Sample_Size_M'].append(n_undergrad)
    results['U_Statistic'].append(u_stat)
    results['P_Value'].append(p_val)
    # results['Preferred_Group'].append(preferred)

# Output DataFrame and save
results_df = pd.DataFrame(results)
# results['Preferred_Group'].append(preferred)
results_df = results_df.sort_values(by='U_Statistic', ascending=False)
results_df.to_csv('/Users/prazeina/Documents/repos/RS/2025/mann-whitney/result/gender_rank_mannWhitneyU_results_all.csv', index=False)

# ---------------------------------------------------------------

header_row1 = ['', 'Female', '', 'Male',  '', '']
header_row2 = ['Skill', 'Median', 'N', 'Median', 'N',  'P-Value']

# Create the new DataFrame with the necessary data columns
output_df = pd.DataFrame({
    'Skill': results_df['Skill'],
    'Female_Median': results_df['Median_F'],
    'Female_N': results_df['Sample_Size_F'],
    'Male_Median': results_df['Median_M'],
    'Male_N': results_df['Sample_Size_M'],
    # 'U_Statistic': results_df['U_Statistic'],
    'P-Value': results_df['P_Value']
})

# --- Save to Excel ---
# --- Sort by U_Statistic in descending order ---
# output_df = output_df.sort_values(by='U_Statistic', ascending=False)

wb = Workbook()
ws = wb.active
ws.title = "Results"

# Write the header rows
ws.append(header_row1)
ws.append(header_row2)

# Write the data rows
for row in dataframe_to_rows(output_df, index=False, header=False):
    ws.append(row)

# Merging cells under "Female" and "Maleuate" headers
ws.merge_cells('B1:C1')  # Merge Female columns (Median and N)
ws.merge_cells('D1:E1')  # Merge Maleuate columns (Median and N)

# Apply center alignment for merged cells
ws['B1'].alignment = Alignment(horizontal='center')
ws['D1'].alignment = Alignment(horizontal='center')

# Save the Excel file
excel_output_file_path = '/Users/prazeina/Documents/repos/RS/2025/mann-whitney/result/gender_rank_mannWhitneyU_results.xlsx'
wb.save(excel_output_file_path)
# --------------------------------------------------------------------------------
import pandas as pd
import dataframe_image as dfi

# Read the Excel file, specifying no header initially to preserve the multi-level structure
import pandas as pd
import dataframe_image as dfi

# Read the Excel file
df = pd.read_excel(
    "/Users/prazeina/Documents/repos/RS/2025/mann-whitney/result/gender_rank_mannWhitneyU_results.xlsx",
    sheet_name="Results",
    header=None
)

# Set the first two rows as MultiIndex header
df.columns = pd.MultiIndex.from_arrays([df.iloc[0], df.iloc[1]])
df = df.drop([0, 1]).reset_index(drop=True)

# Assign clean column names
new_columns = pd.MultiIndex.from_tuples([
    ('', 'Skill'),
    ('Female', 'Median'),
    ('Female', 'N'),
    ('Male', 'Median'),
    ('Male', 'N'),
    # ('', 'U-Statistic'),
    ('', 'p-value')
])
df.columns = new_columns

# Function to apply border styling based on p-value
def highlight_pval(val):
    try:
        val = float(val)
        if val < 0.05:
            return 'border: 2px solid red'
        elif val < 0.1:
            return 'border: 2px solid green'
        else:
            return ''
    except:
        return ''

df_styled = df.style \
    .hide(axis='index') \
    .format({
        ('', 'p-value'): lambda x: f"{x:.3f}".lstrip("0"),
        # ('', 'U-Statistic'): "{:.0f}",
        ('Female', 'Median'): lambda x: f"{float(x):.1f}".rstrip("0").rstrip(".").lstrip("0") if "." in f"{float(x):.1f}" else f"{int(x)}",
        ('Female', 'N'): "{:.0f}",
        ('Male', 'Median'): lambda x: f"{float(x):.1f}".rstrip("0").rstrip(".").lstrip("0") if "." in f"{float(x):.1f}" else f"{int(x)}",
        ('Male', 'N'): "{:.0f}"
    }) \
    .applymap(highlight_pval, subset=[('', 'p-value')]) \
    .set_table_styles([
        {'selector': 'tbody tr:nth-child(even)', 'props': [('background-color', '#f9f9f9')]}
    ])

# Export as image
dfi.export(df_styled, "/Users/prazeina/Documents/repos/RS/2025/mann-whitney/result/2025 Rank Gender.png", dpi=300)

# import matplotlib.pyplot as plt

# # Load the dataset
# df = pd.read_csv("/Users/prazeina/Documents/repos/RS/2025/mann-whitney/result/gender_rank_mannWhitneyU_results.csv")

# # Extract relevant columns for plotting
# skills = df['Skill']
# mean_graduates = df['Mean_F']
# mean_undergraduates = df['Mean_M']

# # Plotting
# plt.figure(figsize=(10, 6))

# # Create two bars for each skill: one for Graduates, one for Undergraduates
# bar_width = 0.35  # Width of the bars
# index = range(len(skills))  # Position of the bars

# plt.barh([i - bar_width / 2 for i in index], mean_graduates, bar_width, label='F', color='#00386C')
# plt.barh([i + bar_width / 2 for i in index], mean_undergraduates, bar_width, label='M', color='#FFC333')

# # Add labels and title
# plt.xlabel('Weighted Average Rank')
# plt.ylabel('Skill')
# plt.title('2025 - Weighted Mean Rank of Skills by Gender\n(Lower = More Important)')
# plt.yticks(index, skills)
# plt.legend(loc='lower right')

# # Show the plot
# plt.tight_layout()
# plt.show()

